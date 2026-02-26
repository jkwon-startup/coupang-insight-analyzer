"""Excel 내보내기 (3시트: 스토리/리뷰/문의)"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class ExcelExporter:
    def generate(
        self,
        product_data,
        reviews,
        qna_pairs,
        story_result,
        review_result,
        qna_result,
        full_result,
    ) -> bytes:
        wb = Workbook()

        # Sheet 1: 스토리 분석
        ws1 = wb.active
        ws1.title = "스토리 분석"
        self._write_story_sheet(ws1, product_data, story_result, full_result)

        # Sheet 2: 리뷰 분석
        ws2 = wb.create_sheet("리뷰 분석")
        self._write_review_sheet(ws2, reviews, review_result)

        # Sheet 3: Q&A 분석
        ws3 = wb.create_sheet("문의 분석")
        self._write_qna_sheet(ws3, qna_pairs, qna_result)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _write_story_sheet(self, ws, product_data, story_result, full_result):
        # 상품 정보
        ws.append(["상품 분석 리포트"])
        ws.merge_cells("A1:D1")
        ws["A1"].font = Font(bold=True, size=14)

        row = 3
        if product_data:
            info = [
                ("상품명", product_data.get("title", "")),
                ("가격", product_data.get("price", "")),
                ("리뷰 수", str(product_data.get("review_count", ""))),
                ("URL", product_data.get("url", "")),
            ]
            for label, value in info:
                ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row, column=2, value=value)
                row += 1

        row += 1
        ws.cell(row=row, column=1, value="스토리 플로우 분석").font = Font(bold=True, size=12)
        row += 1
        if story_result:
            for line in story_result.split("\n"):
                ws.cell(row=row, column=1, value=line)
                row += 1

        row += 2
        if full_result:
            ws.cell(row=row, column=1, value="종합 리포트").font = Font(bold=True, size=12)
            row += 1
            for line in full_result.split("\n"):
                ws.cell(row=row, column=1, value=line)
                row += 1

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 60

    def _write_review_sheet(self, ws, reviews, review_result):
        # 리뷰 원본 데이터 테이블
        headers = ["번호", "별점", "작성자", "날짜", "내용", "도움"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        for i, r in enumerate(reviews, 1):
            ws.cell(row=i + 1, column=1, value=i).border = THIN_BORDER
            ws.cell(row=i + 1, column=2, value=r.get("rating")).border = THIN_BORDER
            ws.cell(row=i + 1, column=3, value=r.get("author", "")).border = THIN_BORDER
            ws.cell(row=i + 1, column=4, value=r.get("date", "")).border = THIN_BORDER
            content = r.get("content", "")
            if r.get("headline"):
                content = f"[{r['headline']}] {content}"
            ws.cell(row=i + 1, column=5, value=content).border = THIN_BORDER
            ws.cell(row=i + 1, column=5).alignment = CELL_ALIGNMENT
            ws.cell(row=i + 1, column=6, value=r.get("helpful", 0)).border = THIN_BORDER

        # AI 분석 결과
        result_row = len(reviews) + 4
        ws.cell(row=result_row, column=1, value="AI 리뷰 분석 결과").font = Font(bold=True, size=12)
        result_row += 1
        if review_result:
            for line in review_result.split("\n"):
                ws.cell(row=result_row, column=1, value=line)
                result_row += 1

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 60
        ws.column_dimensions["F"].width = 8

    def _write_qna_sheet(self, ws, qna_pairs, qna_result):
        headers = ["번호", "질문", "답변", "질문일", "판매자"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        for i, q in enumerate(qna_pairs, 1):
            ws.cell(row=i + 1, column=1, value=i).border = THIN_BORDER
            ws.cell(row=i + 1, column=2, value=q.get("question", "")).border = THIN_BORDER
            ws.cell(row=i + 1, column=2).alignment = CELL_ALIGNMENT
            ws.cell(row=i + 1, column=3, value=q.get("answer", "")).border = THIN_BORDER
            ws.cell(row=i + 1, column=3).alignment = CELL_ALIGNMENT
            ws.cell(row=i + 1, column=4, value=q.get("q_date", "")).border = THIN_BORDER
            ws.cell(row=i + 1, column=5, value=q.get("seller", "")).border = THIN_BORDER

        result_row = len(qna_pairs) + 4
        ws.cell(row=result_row, column=1, value="AI Q&A 분석 결과").font = Font(bold=True, size=12)
        result_row += 1
        if qna_result:
            for line in qna_result.split("\n"):
                ws.cell(row=result_row, column=1, value=line)
                result_row += 1

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 15
